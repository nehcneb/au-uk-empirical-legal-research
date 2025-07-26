# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.1
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Preliminaries

# %%
def own_account_allowed():
    return 0


# %%
def check_questions_answers():
    return 1


# %%
def batch_mode_allowed():
    return 1


# %%
def pop_judgment():
    return 1


# %%
def immediate_b64():
    return 0


# %%
#Default judgment counter bound
default_judgment_counter_bound = 10

#Cutoff for requiring batch mode
judgment_batch_cutoff = 25 #Change this at home

#max number of judgments under batch mode
judgment_batch_max = 200

# %%
#Default page bound for OWN.py
default_page_bound = 100

# %%
huggingface = True

#if depends on director
huggingface_directory = 0

if huggingface_directory > 1:
    
    current_dir = ''
    
    try:
        current_dir = os.getcwd()
        print(f"current_dir == {current_dir}")
        
    except Exception as e:
        print(f"current_dir not generated.")
        print(e)
    
    if 'Users/Ben' not in current_dir: #If running on Huggingface or Github Actions
        huggingface = True
        
print(f'huggingface == {huggingface}')


# %%
#Preliminaries
import datetime
from datetime import date
from dateutil import parser
#from dateutil.relativedelta import *
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import requests
import pypdf
import os
import io
from io import BytesIO
from io import StringIO
import pause
import re
import mammoth

#PDF images
import pdf2image
from PIL import Image
import pytesseract

#Excel
import openpyxl
from pyxlsb import open_workbook as open_xlsb
import xlsxwriter

#Streamlit
import streamlit as st
#from streamlit_gsheets import GSheetsConnection
from streamlit.components.v1 import html
#import streamlit_ext as ste

#AWS
import boto3
from botocore.config import Config
from botocore.exceptions import ClientError


# %% [markdown]
# # Scraper etc

# %%
#Check if string is date

#From https://stackoverflow.com/questions/25341945/check-if-string-has-date-any-format

def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try: 
        parser.parse(string, fuzzy=fuzzy)
        return True

    except ValueError:
        return False



# %%
def date_parser(x):

    #st.write(f"x = {x}")
    
    try:

        if isinstance(x, datetime):
            
            return x

        else:
            
            #Determine if day first or year first
            first_digits_list = re.findall(r'^\d+', str(x))

            if len(first_digits_list) > 0:

                first_digits = first_digits_list[0]

                #st.write(f"first_digits == {first_digits}")
                
                if len(first_digits) == 4:
                
                    return parser.parse(str(x), yearfirst=True)
        
                elif len(first_digits) in [1, 2]:
        
                    return parser.parse(str(x), dayfirst=True)
        
                else:
        
                    return None
            else:
                return None
    
    except Exception as e:

        #st.write(e)
        
        return None
        


# %%
#today
today_in_nums = str(datetime.now())[0:10]
today = datetime.now().strftime("%d/%m/%Y")

# %%
# Generate placeholder list of errors
errors_list = set()


# %%
#Split title and mnc from full case title

def split_title_mnc(full_title):
    
    #Returns a list where first item is full_title and second item mnc

    full_title = str(full_title) #This is to convert any nan float to str
    
    #Get rid of extra spaces
    while '  ' in full_title:
        
        full_title = full_title.replace('  ', ' ')

    #Get mnc
    mnc = ''
    
    mnc_list = re.findall(r'\[\d{4}\]\s\D+\s\d+', full_title)

    if len(mnc_list) > 0:
        
        mnc = mnc_list[0]
        
        if isinstance(mnc, tuple):
            
            mnc = mnc[0]
    
    #Get title 
    if (len(mnc) > 0) and (mnc in full_title):
        
        title = full_title.split(mnc)[0]

        if len(title) > 0:
            
            while title[0] == ' ':
                title = title[1:]

            while title[-1] == ' ':
                title = title[:-1]
                
    else:
        title = full_title
    
    return [title, mnc]



# %%
#Pause between judgment scraping
scraper_pause_mean = int(10)


# %%
#Lowerbound on length of judgment text to proccess without trying to download directly from official database, in tokens
judgment_text_lower_bound = 4000 #~3000 words


# %%
#Tidy up hyperlink
def link(x):
    
    if len(str(x)) > 0:
        
        value = '=HYPERLINK("' + str(x) + '")'
        
        return value
    
    else:
        
        return x
        


# %%
#Define function for judgment link containing PDF
def pdf_judgment(url):
    headers = {'User-Agent': 'whatever'}

    #print(url)
    
    r = requests.get(url, headers=headers, allow_redirects = True)

    #print('Got judgment bytes data')
    
    remote_file_bytes = io.BytesIO(r.content)
    pdfdoc_remote = pypdf.PdfReader(remote_file_bytes)
    text_list = []

    for page in pdfdoc_remote.pages:
        text_list.append(page.extract_text())
    
    return str(text_list)



# %%
#Define function for judgment link containing PDF images
def pdf_image_judgment(url):
    headers = {'User-Agent': 'whatever'}

    r = requests.get(url, headers=headers, allow_redirects = True)
    
    remote_file_bytes = r.content

    images = pdf2image.convert_from_bytes(remote_file_bytes, timeout=30)
    
    #Extract text from images
    text_list = []
    
    max_images_number = len(images)

    for image in images[ : max_images_number]:
        
        text_page = pytesseract.image_to_string(image, timeout=30)
        
        text_list.append(text_page)
        
    return str(text_list)



# %%
#Define function for judgment link containing docx
def docx_judgment(url):
    headers = {'User-Agent': 'whatever'}
    r = requests.get(url, headers=headers)
    remote_file_bytes = io.BytesIO(r.content)
    
    doc_string = mammoth.convert_to_html(remote_file_bytes).value
    
    return doc_string



# %%
def tips():
    st.markdown(""":green[**DO's**:]
- :green[Do break down complex tasks into simple sub-tasks.]
- :green[Do give clear and detailed instructions (eg specify steps required to complete a task).]
- :green[Do use the same terminology as the relevant judgments or files themselves.]
- :green[Do give exemplar answers.]
- :green[Do manually check some or all answers.]
- :green[Do revise questions to get better answers.]
- :green[Do evaluate answers on the same sample of judgments or files (ie the "training" sample).]
""")

    st.markdown(""":red[**Don'ts**:]
- :red[Don't ask questions which go beyond the relevant judgments or files themselves.]
- :red[Don't ask difficult maths questions.]
- :red[Don't skip manual evaluation.]
""")

    st.markdown(""":orange[**Maybe's**:]
- :orange[Maybe ask for reasoning.]
- :orange[Maybe re-run the same questions and manually check for inconsistency.]
""")

    st.write('Click [here](https://platform.openai.com/docs/guides/prompt-engineering) for more tips.')


# %%
def list_value_check(some_list, some_value):
    try:
        index = some_list.index(some_value)
        return index
    except:
        return None
        


# %%
#Function for changing selection menu for type on Streamlit

def dict_value_or_none(some_dict, some_key):

    if (some_key in [None, '']) or (not isinstance(some_dict, dict)):

        return None
    
    elif some_key not in some_dict.keys():
        
        return None
    
    else:

        return_value = some_dict[some_key]

        if isinstance(return_value, dict):
            
            return_value = [*return_value.keys()]
        
        return return_value
    


# %%
def list_range_check(some_list, some_string):
    selected_list = []
    try:
        raw_list = some_string.split(',')

        for item in raw_list:

            while item[0] == ' ':
                item = item[1:]
            
            if item in some_list:
                selected_list.append(item)

    except:
        if ((some_string != '') or (some_string != None)):
            print(f'List {str(some_list)} does not contain {some_string}')
 
    return selected_list



# %%
#Function for turning month or year choice to number or empty string

def month_year_to_str(x):

    if not re.search(r'\d+', str(x)):

        return ''

    else:
        
        return re.findall(r'\d+', str(x))[0]



# %%
#String to integer
def str_to_int(string):
    try:
        output = int(float(string))
        return output
    except:
        return 1


# %%
#String to integer
def str_to_int_page(string):
    try:
        output = int(float(string))
        return output
    except:
        return int(default_page_bound)


# %%
#Keys to keep while switching pages
keys_to_carry_over = ['Your name', 
                    'Your email address', 
                    'Your GPT API key', 
                    'Maximum number of judgments', 
                    'Maximum number of files',
                    'Maximum number of pages per file',
                    'Language choice',
                    'Enter your questions for GPT', 
                    'Use GPT', 
                    'Use own account', 
                    'Use flagship version of GPT', 
                    'System',
                    'Consent', 
                     'submission_time', 
                     'status', 
                      'jurisdiction_page', 
                      #'CourtListener API token' #US specific
                     ]


# %%
#Save jurisdiction specific input
def save_input(df_master):
    
    #df_master = df_master.replace({np.nan: None})
    
    for key in st.session_state.df_master.keys():
    
        if (key not in keys_to_carry_over) and key in df_master.columns:
            
            try:
                
                st.session_state.df_master.loc[0, key]  = df_master.loc[0, key]
                
            except Exception as e:
                
                print(f"{key} of {type(df_master.loc[0, key])} not saved, trying to convert type of st.session_state.df_master[{key}] to 'object'.")

                if isinstance(df_master.loc[0, key], list):
                
                    st.session_state.df_master[key] = st.session_state.df_master[key].astype('object')
                    
                    st.session_state.df_master.at[0, key] = df_master.loc[0, key]

                    print(f"{key} of {type(df_master.loc[0, key])} now saved.")

            except Exception as e2:
                
                print(f"{key} still not saved after converting type of st.session_state.df_master[{key}] to 'object'.")

                print(e2)



# %%
#Function to hide own token

def hide_own_token(user_token, own_token):
    if user_token:
        if user_token == own_token:
            return None
        else:
            return user_token
    else:
        return None



# %%
#Reverse hyperlink display
def reverse_link(x):
    value = str(x).replace('=HYPERLINK("', '').replace('")', '')
    return value


# %%
#Display error for scraping
#search_error_display = 'The database from which this app sources cases is not responding. Please try again in a few hours.'
search_error_display = 'Sorry, an error has occurred. Please change your entries or wait a few hours, and try again.'


# %%
#Note error for scraping
search_error_note = 'Error: case/file not sent to GPT given full text was not scrapped.'


# %%
#Note truncation
truncation_note = "The full file is too long for GPT. It was (or will be) truncated if sent to GPT."

# %%
# Programmaticaly produced GPT headings
own_gpt_headings = ['Hyperlink', 'in tokens (up to', 'GPT cost estimate', 'GPT time estimate']


# %% [markdown]
# # Streamlit

# %%
#Function to open url
def open_page(url):
    open_script= """
        <script type="text/javascript">
            window.open('%s', '_blank').focus();
        </script>
    """ % (url)
    html(open_script)


# %%
def clear_cache():
    
    keys = list(st.session_state.keys())
    
    for key in keys:
        st.session_state.pop(key)



# %%
def clear_cache_except_validation_df_master():
    keys = list(st.session_state.keys())
    if 'df_master' in keys:
        keys.remove('df_master')
    if 'page_from' in keys:
        keys.remove('page_from')
    if 'jurisdiction_page' in keys:
        keys.remove('jurisdiction_page')
    for key in keys:
        st.session_state.pop(key)


# %%
def streamlit_timezone():
    local_now = datetime.now().astimezone()
    time_zone = local_now.tzname()
    if time_zone in ['AEST', 'AEDT', 'ACST', 'AWST', 'BST']:
        return True
    else:
        return False


# %%
def streamlit_cloud_date_format(date):
    local_now = datetime.now().astimezone()
    time_zone = local_now.tzname()
    if time_zone in ['AEST', 'ACST', 'AWST', 'BST']:
        date_to_send = parser.parse(date, dayfirst=True).strftime("%d/%m/%Y")
    else:
        date_to_send = parser.parse(date, dayfirst=False).strftime("%m/%d/%Y")
    return date_to_send


# %%
#No results msg
no_results_msg = 'Your search terms returned 0 results. Please change your search terms and try again. \nIf this problem persists, please close this app and try again later.'

# %%
#Default spinner_text

#spinner_text = r"$\textsf{\normalsize In progress... }$"
spinner_text = 'In progress...'


# %%
#Funder
#funder_msg = "Lawtodata is partially funded by a 2022 University of Sydney Research Accelerator (SOAR) Prize and a 2023 Discovery Early Career Researcher Award (DECRA). Please kindly acknowledge this if you use your requested data to produce any research output. "

funder_msg = "I developed LawtoData with a view to promoting and advancing quantitative legal research. Please feel free to share this app with others if you find it useful. "


# %%
#Cost
gpt_cost_msg = "This app uses a costly GPT service. For the default model, Ben's own experience suggests that it costs approximately USD \$0.01 (excl GST) per file. The [exact cost](https://openai.com/api/pricing/) for answering a question about a file depends on the length of the question, the length of the file, and the length of the answer produced. You will be given ex-post cost estimates."


# %%
#Create function for saving responses and results
def convert_df_to_json(df):
    return df.to_json(orient = 'split', compression = 'infer', default_handler=str, indent=4)

def convert_df_to_csv(df):
   return df.to_csv(index=False).encode('utf-8')

def convert_df_to_excel(df):
    #Excel metadata
    excel_author = 'LawtoData'
    excel_description = 'A 2022 University of Sydney Research Accelerator (SOAR) Prize and a 2023 Discovery Early Career Researcher Award (DECRA) partially funded the development of LawtoData, which generated this spreadsheet.'
    output = BytesIO()
    #writer = pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_urls': False}})
    writer = pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_urls': False}})
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    workbook.set_properties({"author": excel_author, "comments": excel_description})
    worksheet = writer.sheets['Sheet1']
#    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None)#, format1)  
    #writer.save()
    writer._save()
    processed_data = output.getvalue()
    return processed_data


# %%
#Download entries and results
@st.fragment
def download_buttons(df_master, df_individual = [], saving = False, previous = False):
    #Enable the saving argument if want to allow saving of entries
    #Enable the previous argument if want to allow saving of last produced results
    #Default df_individual is empty to ensure no buttons for downloading data is shown
    
    #Create a copy of df_master to avoid exposing secrets
    df_master_to_show = df_master.copy(deep = True)

    #For downloading entries
    if saving:

        if 'Your GPT API key' in df_master_to_show.columns:
            df_master_to_show["Your GPT API key"] = ''

        if 'CourtListener API token' in df_master_to_show.columns:
            #Essential to avoiding both default and user secrets
            df_master_to_show["CourtListener API token"] = ''

        if previous:
            
            st.warning('Looking for your entries?')
            
            previous_str = 'your last produced '
            
        else:
            
            st.info('Your entries are now available for download.')
        
        responses_output_name = str(df_master_to_show.loc[0, 'Your name']) + '_' + str(today_in_nums) + '_entries'
        
        xlsx = convert_df_to_excel(df_master_to_show)
        
        st.download_button(label='DOWNLOAD your entries as an Excel spreadsheet (XLSX)',
                            data=xlsx,
                            file_name=responses_output_name + '.xlsx', 
                            mime='application/vnd.ms-excel',
                           )
    
        csv = convert_df_to_csv(df_master_to_show)
    
        st.download_button(
            label="DOWNLOAD your entries as a CSV", 
            data = csv,
            file_name=responses_output_name + '.csv', 
            mime= "text/csv", 
        )
    
        json = convert_df_to_json(df_master_to_show)
        
        st.download_button(
            label="DOWNLOAD your entries as a JSON", 
            data = json,
            file_name= responses_output_name + '.json', 
            mime= "application/json", 
        )
    
    #For downloading data
    if len(df_individual) > 0:

        #Determine whether to note previous results
        previous_str = ''
        
        if previous:
            st.info('Looking for your last produced data?')
            
            previous_str = 'last produced '
            
        else:
            
            st.success("Your data is now available for download. Thank you for using *LawtoData*!")

        #Produce output spreadsheets
        output_name = str(df_master_to_show.loc[0, 'Your name']) + '_' + str(today_in_nums) + '_output'

        excel_xlsx = convert_df_to_excel(df_individual)
        
        st.download_button(label=f'DOWNLOAD your {previous_str}data as an Excel spreadsheet (XLSX)',
                            data=excel_xlsx,
                            file_name= output_name + '.xlsx', 
                            mime='application/vnd.ms-excel',
                           )
    
        csv_output = convert_df_to_csv(df_individual)
        
        st.download_button(
            label=f'DOWNLOAD your {previous_str}data as a CSV', 
            data = csv_output,
            file_name= output_name + '.csv', 
            mime= "text/csv", 
        )
        
        json_output = convert_df_to_json(df_individual)
        
        st.download_button(
            label=f'DOWNLOAD your {previous_str}data as a JSON', 
            data = json_output,
            file_name= output_name + '.json', 
            mime= "application/json", 
        )
    
        st.page_link('pages/AI.py', label=f"ANALYSE your {previous_str}data with an AI", icon = '🤔')

    #For noting a lack of data
    if ((not saving) and (len(df_individual) == 0)):
        
        st.error('Sorry, no data was produced. Please return to the previous page, check your search terms and try again.')



# %%
#Obtain columns with hyperlinks

def link_headings_picker(df):
    link_headings = []
    for heading in df.columns:
        if (('hyperlink' in str(heading).lower()) or (str(heading).lower() == 'uri') or (str(heading).lower() == 'url')):
            link_headings.append(heading)
    return link_headings #A list of headings with hyperlinks

def clean_link_columns(df):
        
    link_headers_list = link_headings_picker(df)

    for link_header in link_headers_list:
        if 'hyperlink' in str(link_header).lower():
            df[link_header] = df[link_header].apply(reverse_link)
        
    return df
    


# %%
#Function for preparing df for display with clickable links

def display_df(df):

    #Obtain clolumns with hyperlinks
    
    link_heading_config = {} 
    
    link_headings_list = link_headings_picker(df)
            
    for link_heading in link_headings_list:

        link_heading_config[link_heading] = st.column_config.LinkColumn()

        #If want to display "click" instead of the whole link. Problem with this: if there is an empty cell, 'click' will still be dipslayed.
        #link_heading_config[link_heading] = st.column_config.LinkColumn(display_text = 'Click')
        
        link_heading_config[link_heading] = st.column_config.LinkColumn()

    #Reverse columns with clickable links to raw uri
    df = clean_link_columns(df)

    return {'df': df, 'link_heading_config': link_heading_config}
    


# %%
#For checking whether date entered is within allowable range

def date_range_check(date_start, date_end, date_entry):
    #All arguments are datetime objects

    try:
        if ((date_start <= date_entry) and (date_entry <= date_end)):
            return date_entry
        else:
            return None
    except:
        return None


# %%
#Excel to df with hyperlinks

def excel_to_df_w_links(uploaded_file):

    df = pd.read_excel(uploaded_file)
    
    wb = openpyxl.load_workbook(uploaded_file)
    
    sheets = wb.sheetnames
    
    ws = wb[sheets[0]]

    columns_w_links = link_headings_picker(df)

    for column in columns_w_links:
        
        column_index = list(df.columns).index(column) + 1 #Adding 1 because excel column starts with 1 not 0
        
        row_length = len(df)

        for row in range(0, row_length):
            
            row_index = row + 2 #Adding 1 because excel non-heading row starts with 2 while pandas at 0
            
            try:
	            new_cell = ws.cell(row=row_index, column=column_index).hyperlink.target
            
            except:

	            new_cell = (str(ws.cell(row=row_index, column=column_index).value))
            
            df.loc[row, column] = new_cell
            
    return df


# %%
#Function for getting a df from an uploaded spreadsheet
def uploaded_file_to_df(uploaded_file):
    
    #Get uploaded file extension
    extension = uploaded_file.name.split('.')[-1].lower()
    
    if extension == 'csv':
        df = pd.read_csv(uploaded_file)

    if extension == 'xlsx':
        
        #df = pd.read_excel(uploaded_file)
        
        df = excel_to_df_w_links(uploaded_file)

    if extension == 'json':
        
        df = pd.read_json(uploaded_file)

    return df
    


# %%
#Function for reporting error

def report_error(error_msg, jurisdiction_page, df_master):

    #Send me an email to let me know
    report_error_email(ULTIMATE_RECIPIENT_NAME = df_master.loc[0, 'Your name'], 
                            ULTIMATE_RECIPIENT_EMAIL = df_master.loc[0, 'Your email address'],
                       jurisdiction_page = jurisdiction_page,
                       df_master = df_master, 
                       error_msg = error_msg
                        )

    st.success("Thank you for reporting the error. We will look at your report as soon as possible.")
    
    #Clear any error
    error_msg_to_return = ''
    
    return error_msg_to_return



# %% [markdown]
# # AWS

# %%
#Get credentials

#If running on Github Actions, then '/home/runner/' in current_dir

#Try local or streamlit first

try:
    
    API_key = st.secrets["openai"]["gpt_api_key"]
    
    AWS_DEFAULT_REGION=st.secrets["aws"]["AWS_DEFAULT_REGION"]
    AWS_ACCESS_KEY_ID=st.secrets["aws"]["AWS_ACCESS_KEY_ID"]
    AWS_SECRET_ACCESS_KEY=st.secrets["aws"]["AWS_SECRET_ACCESS_KEY"]
    
    SENDER = st.secrets["email_notifications"]["email_sender"]
    RECIPIENT = st.secrets["email_notifications"]["email_receiver_work"]
        
    print('Running locally or on Streamlit')
    
except:
    
    API_key = os.environ['GPT_API_KEY']
    
    AWS_DEFAULT_REGION = os.environ['AWS_DEFAULT_REGION']
    AWS_ACCESS_KEY_ID = os.environ['AWS_ACCESS_KEY_ID']
    AWS_SECRET_ACCESS_KEY = os.environ['AWS_SECRET_ACCESS_KEY']
    
    SENDER = os.environ['EMAIL_SENDER']
    RECIPIENT = os.environ['EMAIL_RECEIVER_WORK']

    print('Running on GitHub Actions or HuggingFace')


# %%
#Function for initiating aws s3

def get_aws_s3():
    
    #Initiate aws s3
    s3_resource = boto3.resource('s3', region_name = AWS_DEFAULT_REGION, aws_access_key_id = AWS_ACCESS_KEY_ID, aws_secret_access_key = AWS_SECRET_ACCESS_KEY)

    return s3_resource



# %%
#Function for getting df from aws
def aws_df_get(s3_resource, df_name):
#df_name is a string of the file name of the relevant df to get from aws, WITH the extension (ie csv)
#Returns the relevant df as Pandas object if found, or an empty Pandas object if not found or other error

    try:

        #s3_resource = get_aws_s3()
        
        #Get relevant df from aws
        obj = s3_resource.Object('lawtodata', df_name).get()
        body = obj['Body'].read()

        df = pd.read_csv(BytesIO(body), index_col=0)
        
        print(f"Sucessfully loaded {df_name} from aws.")

    except Exception as e:

        print(f"Failed to load {df_name} from aws due to error: {e}.")

        df = pd.DataFrame([])
        
    return df



# %%
#Function for uploading df to aws
def aws_df_put(s3_resource, df, df_name):
#df is the Pandas object
#df_name is a string of the file name of the relevant df to upload to aws, WITH the extension (ie csv)

    csv_buffer = StringIO()
    df.to_csv(csv_buffer)
    s3_resource.Object('lawtodata', df_name).put(Body=csv_buffer.getvalue())
    


# %%
#Get all objects from aws s3
#NOT IN USE

def get_aws_objects():
    
    #Get a list of all files on s3
    bucket = s3_resource.Bucket('lawtodata')
    
    aws_objects = []
    
    for obj in bucket.objects.all():
        key = obj.key
        body = obj.get()['Body'].read()
        key_body = {'key': key, 'body': body}
        aws_objects.append(key_body)

    return aws_objects



# %%
#Function for using aws ses for sending emails
def get_aws_ses():
    ses = boto3.client('ses',region_name = AWS_DEFAULT_REGION, aws_access_key_id = AWS_ACCESS_KEY_ID, aws_secret_access_key = AWS_SECRET_ACCESS_KEY)
    #ses is based on the following upon substitutiong 'ses' for 's3', https://boto3.amazonaws.com/v1/documentation/api/latest/guide/credentials.html#guide-credentials
    return ses


# %%
#AWS email
#Define send email function

def send_notification_email(ULTIMATE_RECIPIENT_NAME, ULTIMATE_RECIPIENT_EMAIL, jurisdiction_page):

    #ses = boto3.client('ses',region_name = AWS_DEFAULT_REGION, aws_access_key_id = AWS_ACCESS_KEY_ID, aws_secret_access_key = AWS_SECRET_ACCESS_KEY)
    ses = get_aws_ses()
    
    #Based on the following upon substituting various arguments, https://docs.aws.amazon.com/ses/latest/dg/send-an-email-using-sdk-programmatically.html
    
    # Replace sender@example.com with your "From" address.
    # This address must be verified with Amazon SES.
    SENDER = st.secrets["email_notifications"]["email_sender"]
    
    # Replace recipient@example.com with a "To" address. If your account 
    # is still in the sandbox, this address must be verified.
    RECIPIENT = st.secrets["email_notifications"]["email_receiver_work"]
    
    # The subject line for the email.
    SUBJECT = f"LawtoData: {ULTIMATE_RECIPIENT_NAME} has requested data"
    
    BODY_TEXT = (
    
    f"{ULTIMATE_RECIPIENT_NAME} at {ULTIMATE_RECIPIENT_EMAIL} has requested data for {jurisdiction_page} via LawtoData."
    
    )
      
    # The character encoding for the email.
    CHARSET = "UTF-8"

    # Try to send the email.
    try:
        #Provide the contents of the email.
        response = ses.send_email(
            Destination={
                'ToAddresses': [
                    RECIPIENT,
                ],
            },
            Message={
                'Body': {
                    'Text': {
                        'Charset': CHARSET,
                        'Data': BODY_TEXT,
                    },
                },
                'Subject': {
                    'Charset': CHARSET,
                    'Data': SUBJECT,
                },
            },
            Source=SENDER,
        )
    # Display an error if something goes wrong.	
    except ClientError as e:
        print(e.response['Error']['Message'])
    #else:
        #print("Email sent! Message ID:"),
        #print(response['MessageId'])


# %%
#AWS email
#Define report error email function

def report_error_email(ULTIMATE_RECIPIENT_NAME, ULTIMATE_RECIPIENT_EMAIL, jurisdiction_page, df_master, error_msg):

    #ses = boto3.client('ses',region_name = AWS_DEFAULT_REGION, aws_access_key_id = AWS_ACCESS_KEY_ID, aws_secret_access_key = AWS_SECRET_ACCESS_KEY)
    ses = get_aws_ses()
    
    #Based on the following upon substituting various arguments, https://docs.aws.amazon.com/ses/latest/dg/send-an-email-using-sdk-programmatically.html
    
    # Replace sender@example.com with your "From" address.
    # This address must be verified with Amazon SES.
    SENDER = st.secrets["email_notifications"]["email_sender"]
    
    # Replace recipient@example.com with a "To" address. If your account 
    # is still in the sandbox, this address must be verified.
    RECIPIENT = st.secrets["email_notifications"]["email_receiver_work"]

    #Entries
    entries_string = f'jurisdiction_page: {jurisdiction_page}\r\n'

    for col in df_master.columns:
        if ('API key' not in col) and ('token' not in col):
            cell = df_master.loc[0, col]
            entries_string += f'{col}:{cell}\r\n'
    
    # The subject line for the email.
    SUBJECT = f"LawtoData: {ULTIMATE_RECIPIENT_NAME} has reported an error"
    
    BODY_TEXT = (
    
    f"{ULTIMATE_RECIPIENT_NAME} at {ULTIMATE_RECIPIENT_EMAIL} has reported an error for LawtoData.\r\n\r\n"

    f"df_master is as follows:\r\n\r\n{entries_string}"

    f"\r\n\r\nError is as follows:\r\n\r\n{error_msg}"

    )
    
    # The character encoding for the email.
    CHARSET = "UTF-8"

    # Try to send the email.
    try:
        #Provide the contents of the email.
        response = ses.send_email(
            Destination={
                'ToAddresses': [
                    RECIPIENT,
                ],
            },
            Message={
                'Body': {
                    'Text': {
                        'Charset': CHARSET,
                        'Data': BODY_TEXT,
                    },
                },
                'Subject': {
                    'Charset': CHARSET,
                    'Data': SUBJECT,
                },
            },
            Source=SENDER,
        )
    # Display an error if something goes wrong.	
    except ClientError as e:
        print(e.response['Error']['Message'])
    #else:
        #print("Email sent! Message ID:"),
        #print(response['MessageId'])

# %% [markdown]
# # [NOT IN USE] Google Sheets

# %%
#Keep record on Google sheet
#Obtain google spreadsheet       
#conn = st.connection("gsheets_nsw", type=GSheetsConnection)
#df_google = conn.read()
#df_google = df_google.fillna('')
#df_google=df_google[df_google["Processed"]!='']
#df_master["Processed"] = datetime.now()
#df_master.pop("Your GPT API key")
#df_to_update = pd.concat([df_google, df_master])
#conn.update(worksheet="CTH", data=df_to_update, )
